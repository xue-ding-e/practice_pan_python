import os
from openpyxl import workbook
from openpyxl.styles import PatternFill, Border, Side

# 处理文件路径
base_dir = os.path.dirname(os.path.abspath(__file__))
target_excel_path = os.path.join(base_dir, "users.xlsx")

# 创建excel且默认会创建一个sheet（名称为Sheet）
wb = workbook.Workbook()

sheet = wb.worksheets[0] # 或 sheet = wb["Sheet"]

# 找到单元格，并修改单元格的内容
cell1 = sheet.cell(1, 1)
cell1.fill = PatternFill("solid", fgColor="99ccff")
cell1.value = "用户姓名"

cell2 = sheet.cell(1, 2)
cell2.fill = PatternFill("solid", fgColor="99ccff")
cell2.border = Border(
    left=Side(style="dashed", color="FFB6C1"),
    right=Side(style="dashed", color="9932CC"),
)
cell2.value = "密码"

cell3 = sheet.cell(1, 3)
cell3.fill = PatternFill("solid", fgColor="99ccff")
cell3.value = "注册时间"

wb.save(target_excel_path)